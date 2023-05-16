# (c) E-kvadrat Consulting & Media, 2021

from argparse import ArgumentParser
from pathlib import Path
from typing import Optional
import openpyxl
from openpyxl.utils import get_column_letter
import pandas as pd
import reader
import c1_parser
from name_variant import NameVariants, NameVariantsDict
import csv_utils


def parse_args():
    argp = ArgumentParser()
    argp.add_argument("template", type=Path, help=".xlsx template file")
    argp.add_argument("--ris", type=Path, help="RIS folder")
    argp.add_argument("--ris_ahci", type=Path, help="RIS_AHCI folder")
    argp.add_argument("--wsd", type=Path, help="Web of Science Documents.csv")
    argp.add_argument("--wsd_q1", type=Path, help="Web of Science DocumentsQ1.csv")
    argp.add_argument("--wsd_q2", type=Path, help="Web of Science DocumentsQ2.csv")
    argp.add_argument("--wsd_q3", type=Path, help="Web of Science DocumentsQ3.csv")
    argp.add_argument("--wsd_q4", type=Path, help="Web of Science DocumentsQ4.csv")
    argp.add_argument(
        "--name_variants", required=True, type=Path, help="name_variant.txt file"
    )
    argp.add_argument("-o", "--output", default="out.xlsx", help="Output .xlsx file")
    return argp.parse_args()


args = parse_args()


def try_get_first_line(dictionary: dict, key: str):
    lines = dictionary.get(key, None)
    if lines is not None:
        return lines[0]


def try_load_ris(path: Optional[Path]):
    df_records = []
    if path:
        for ris_file in path.iterdir():
            print("Reading ", ris_file)
            for ris_record in reader.read(ris_file):
                df_records.append(
                    {
                        "UT": ris_record["UT"][0],
                        "PY": try_get_first_line(ris_record, "PY"),
                        "DT": try_get_first_line(ris_record, "DT"),
                        "PT": try_get_first_line(ris_record, "PT"),
                        "AU": ris_record["AF"],
                        "C1": c1_parser.parse(ris_record["C1"]),
                    },
                )
    return pd.DataFrame.from_records(
        df_records, columns=["UT", "PY", "DT", "PT", "AU", "C1"]
    )


ris_df = try_load_ris(args.ris)
ahci_df = try_load_ris(args.ris_ahci)


def try_load_wsd(path: Optional[Path]):
    wsd_records = []
    if path:
        for row in csv_utils.read_csv_body(args.wsd):
            wsd_records.append(
                {
                    "UT": row["Accession Number"],
                    "DT": row["Document Type"],
                    "PY": row["Publication Date"],
                },
            )
    return pd.DataFrame.from_records(wsd_records, columns=["UT", "PY", "DT"])


wsd_df = try_load_wsd(args.wsd)
wsdq_dfs = [
    try_load_wsd(args.wsd_q1),
    try_load_wsd(args.wsd_q2),
    try_load_wsd(args.wsd_q3),
    try_load_wsd(args.wsd_q4),
]

name_dict = NameVariantsDict(NameVariants.from_file(args.name_variants))
main_name = name_dict.get_main_name()

# Объединяем источники:
# OUTER JOIN ris, ahci, wsd
# LEFT JOIN wsd_q*

ris_df["SRC_ris"] = "RIS"
ahci_df["SRC_ahci"] = "AHCI"
wsd_df["SRC_wsd"] = "WSD"

joined_df = ris_df.merge(ahci_df, how="outer", on="UT", suffixes=("_ris", "_ahci"))
joined_df = joined_df.merge(
    wsd_df.rename(columns={"DT": "DT_wsd", "PY": "PY_wsd"}), how="outer", on="UT"
)

for q in range(0, 4):
    suffix = f"_wsdq{q+1}"
    qdf = wsdq_dfs[q].rename(columns={"PY": "PY" + suffix, "DT": "DT" + suffix})
    qdf["SRC" + suffix] = f"WSD_Q{q+1}"

    joined_df = joined_df.merge(qdf, how="left", on="UT")

joined_df = joined_df[~joined_df["UT"].duplicated()]  # type:ignore

# Объединяет поля с учетов приоритетов
# TODO: проверить приоритет

joined_df["PY"] = (
    joined_df["PY_ahci"]
    .combine_first(joined_df["PY_wsdq1"])
    .combine_first(joined_df["PY_wsdq2"])
    .combine_first(joined_df["PY_wsdq3"])
    .combine_first(joined_df["PY_wsdq4"])
    .combine_first(joined_df["PY_wsd"])
    .combine_first(joined_df["PY_ris"])
)

joined_df["SRC"] = (
    joined_df["SRC_ahci"]
    .combine_first(joined_df["SRC_wsdq1"])
    .combine_first(joined_df["SRC_wsdq2"])
    .combine_first(joined_df["SRC_wsdq3"])
    .combine_first(joined_df["SRC_wsdq4"])
    .combine_first(joined_df["SRC_wsd"])
    .combine_first(joined_df["SRC_ris"])
)

joined_df["AU"] = joined_df["AU_ahci"].combine_first(joined_df["AU_ris"])  # type:ignore
joined_df["C1"] = joined_df["C1_ahci"].combine_first(joined_df["C1_ris"])  # type:ignore
joined_df["PT"] = joined_df["PT_ahci"].combine_first(joined_df["PT_ris"])  # type:ignore

joined_df["DT"] = joined_df["DT_ahci"].combine_first(joined_df["DT_ris"])  # type:ignore
joined_df["DT InCites"] = (
    joined_df["DT_wsdq1"]
    .combine_first(joined_df["DT_wsdq2"])
    .combine_first(joined_df["DT_wsdq3"])
    .combine_first(joined_df["DT_wsdq4"])
    .combine_first(joined_df["DT_wsd"])
)

wb = openpyxl.load_workbook(args.template)
publications = wb["публикации"]
fractions = wb["фракции"]
name_variants_sheet = wb["Варианты названий университета"]

FORMULA_PUBLICATIONS_FRACTION = f'=IFERROR(SUMIFS(фракции[фракции],фракции[org],"={main_name}",фракции[UT],"="&публикации[[#This Row],[UT]]),"")'
FORMULA_FRACTION_FRACTION = "=1/фракции[count_au]/фракции[count_au_aff]"


# Добавить аффиляцию в таблицу фракций и в словарь имен
def append_fractions(ut, au, count_au, count_au_aff, aff):
    org = "NoN"
    name_aff_variant = name_dict.find_matching_pattern(aff)
    if name_aff_variant is not None:
        org = main_name

    fractions.append(
        [
            ut,
            au,
            count_au,
            count_au_aff,
            aff,
            org,
            name_aff_variant,
            FORMULA_FRACTION_FRACTION,
        ]
    )  # ->фракции


# Квартиль в зависимости от источника
def get_quartile(src):
    if src == "AHCI":
        return "AHCI"
    elif src == "WSD_Q1":
        return "Q1"
    elif src == "WSD_Q2":
        return "Q2"
    elif src == "WSD_Q3":
        return "Q3"
    elif src == "WSD_Q4":
        return "Q4"
    else:
        return "n/a"


# Критерий отбора в зависимости от квартили
def get_criteria_name(quartile: str):
    if quartile == "Q1" or quartile == "Q2" or quartile == "AHCI":
        return "Q1 Q2 AHCI"
    else:
        return "Q3 Q4 n/a"


# TODO: есть ли лучший способ определять, если значение object из DataFrame = NaN?
import math


def unwrap_nan(value, default):
    if type(value) == float and math.isnan(value):
        return default()
    return value


for _, row in joined_df.iterrows():
    ut = row["UT"]
    year = row["PY"]
    document_type = row["DT"]
    document_source = row["PT"]
    document_type_wsd = row["DT InCites"]

    quartile = get_quartile(row["SRC"])
    criteria_name = get_criteria_name(quartile)

    publications.append(
        [
            ut,
            year,
            document_type,
            document_type_wsd,
            document_source,
            criteria_name,
            FORMULA_PUBLICATIONS_FRACTION,
            None,
            quartile,
        ]
    )  # ->публикации

    au_list = unwrap_nan(row["AU"], list)
    count_au = len(au_list)
    c1_data = unwrap_nan(row["C1"], dict)

    if type(c1_data) == dict:
        # В C1 данный вида [Автор] Университет
        for au in au_list:  # каждый автор из AF
            c1_affiliated_unis = c1_data.get(au, [])
            count_au_aff = len(c1_affiliated_unis)

            for (
                aff
            ) in c1_affiliated_unis:  # каждый университет из C1, связанный с автором
                append_fractions(ut, au, count_au, count_au_aff, aff)

    else:
        for au in au_list:
            append_fractions(ut, au, count_au, 1, c1_data)

for aff, _ in name_dict.get_items():
    name_variants_sheet.append([aff, main_name])  # ->Варианты названий университета


def update_table(sheet, name):
    table = sheet.tables[name]
    table.ref = "A1:" + get_column_letter(sheet.max_column) + str(sheet.max_row)


update_table(fractions, "фракции")
update_table(publications, "публикации")
update_table(name_variants_sheet, "варианты_названий_университета")
wb.save(args.output)
